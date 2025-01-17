from moviepy.editor import VideoFileClip
import speech_recognition as sr
from pydub import AudioSegment
import whisper

import csv
import math
import os
import sys
import time

from openpyxl import Workbook

from AzureGPTModelConnection import ChatBot


def transcribe_audio_with_whisper(audio_path, model_name="base"):
    # Load Whisper model
    model = whisper.load_model(model_name)
    # Transcribe audio
    result = model.transcribe(audio_path, language="de")
    return result["text"]

def save_to_csv(data: dict, filename: str): #, encoding='utf-8'):
    # with open(filename, 'w', newline='', encoding=encoding) as csvfile:
    with open(filename, 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=['level_col', 'variable_col'])
        writer.writeheader()
        for key, value in data.items():
            writer.writerow({'level_col': key, 'variable_col': value})

def save_to_excel(data: dict, filename: str):
    wb = Workbook()
    ws = wb.active
    
    # Set header
    ws['A1'] = 'Start time'
    ws['B1'] = 'Text'
    
    # Write data
    row = 2
    for key, value in data.items():
        ws[f'A{row}'] = key
        ws[f'B{row}'] = value
        row += 1
    
    # Save to file
    wb.save(filename)
        
def extract_audio_from_video(video_path, audio_path="extracted_audio.wav"):
    # Load the video file
    video = VideoFileClip(video_path)
    # Extract audio and save it
    video.audio.write_audiofile(audio_path)
    return audio_path

def segment_audio(file_path, segment_length=60000, output_folder="audio_segments"):
    """
    Segments an audio file into chunks.
    
    Parameters:
    - file_path (str): Path to the original audio file.
    - segment_length (int): Length of each segment in milliseconds. (e.g., 60000 ms = 1 minute)
    - output_folder (str): Folder to save the segmented audio files.

    Returns:
    - List of file paths for the audio segments.
    """
    start_time_list = []
    # Load the audio file
    audio = AudioSegment.from_file(file_path)
    
    # Calculate the number of segments
    total_length = len(audio)
    num_segments = math.ceil(total_length / segment_length)
    
    # Ensure output folder exists
    os.makedirs(output_folder, exist_ok=True)
    
    segment_paths = []
    for i in range(num_segments):
        # Calculate start and end times for each segment
        start_time = i * segment_length
        end_time = min((i + 1) * segment_length, total_length)
        
        # Extract segment
        segment = audio[start_time:end_time]
        
        # Export segment to file
        segment_path = os.path.join(output_folder, f"segment_{i+1}.wav")
        segment.export(segment_path, format="wav")
        segment_paths.append(segment_path)
        start_time_list.append(start_time/60000)
        
    return segment_paths, start_time_list

def transcribe_audio(audio_path, language="de-DE"):
    recognizer = sr.Recognizer()
    # Load the audio file
    with sr.AudioFile(audio_path) as source:
        audio_data = recognizer.record(source)
        # Recognize (convert from speech to text)
        try:
            text = recognizer.recognize_google(audio_data, language=language)
        except sr.UnknownValueError:
            text = "Could not understand audio"
        except sr.RequestError:
            text = "Could not request results; check your internet connection"
    return text

def send_request(chat, initial_text, language="English", try_count=0):
    response = chat.send_request_and_get_response(f"Translate the following without any comments into {language}: {initial_text}")
    if isinstance(response, int) and try_count < 5:
        time.sleep(response)
        return send_request(chat, initial_text, try_count + 1)
    else:
        return response

def extract_text_from_video(video_path, video_name, language="English"):
    initial_constructed_str = ''
    initial_text = ''
    item = 0
    time_text_dict = {}
    time_text_dict_eng = {}
    
    chat = ChatBot()
    
    # Step 1: Extract audio from video
    audio_path = extract_audio_from_video(video_path)
    # Step 2: Segment audio to 1 minute
    segments, start_time_list = segment_audio(audio_path, segment_length=60000, output_folder=f"{video_name}") # 60,000 ms = 1 minute
    print("Segments:", segments)
    # Step 3: Transcribe the audio segment with Whisper
    for segment in segments:
        # time.sleep(25)
        text = transcribe_audio_with_whisper(segment)
        initial_text = text.strip() + "\n"
        # print(initial_text)
        response = send_request(chat, initial_text, language)
        # # Handle the response as needed (e.my_string.splitlines()g., print or process)
        data = response.json()
        print(data)
        content = data['choices'][0]['message']['content'] 
        time_text_dict_eng[f"{start_time_list[item]}"] = str(content)
        item+=1

    save_to_excel(time_text_dict_eng, f"{language}_text_{video_name}.csv")
    chat.close_session()


if __name__ == "__main__":
    # An action based on input parameters
    if len(sys.argv) < 2:
        video_path = r"PATH"  # Replace with your video file path
        video_file_name = os.path.basename(video_path)
        transcribed_text = extract_text_from_video(video_path, video_file_name)
    elif len(sys.argv) == 2:
        video_path = sys.argv[1]
        video_file_name = os.path.basename(video_path)
        transcribed_text = extract_text_from_video(video_path, video_file_name)
    elif len(sys.argv) == 3:
        video_path = sys.argv[1]
        video_file_name = os.path.basename(video_path)
        language = sys.argv[2]
        transcribed_text = extract_text_from_video(video_path, video_file_name, language)

